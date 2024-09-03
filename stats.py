import numpy as np
import tkinter as tk
import openpyxl, xlrd 
from openpyxl import Workbook
import pathlib

m = tk.Tk()
m.geometry("800x500")
m.title('Basketball Stats')
m.config(highlightbackground = 'black')

file = pathlib.Path('Backend_Data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1']="Player"
    sheet['B1']="Number"
    sheet['C1']="Points"
    sheet['D1']="Rebounds"
    sheet['E1']="Assists"
    sheet['F1']="Blocks"
    sheet['G1']="Steals"
    sheet['H1']="Turnovers"

    file.save("Backend_Data.xlsx")

def submit():
    a = name.get()
    b= num.get()
    print(a,b)


    file = openpyxl.load_workbook("Backend_Data.xlsx")
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row+1,value=a)
    sheet.cell(column=2, row=sheet.max_row,value=b)

    file.save("Backend_Data.xlsx")

def importPlayers():
    i = tk.Toplevel(m)
    i.geometry("800x500")
    nameLabel = tk.Label(i, text = "Name").place(x=50,y=30)
    numLabel = tk. Label(i, text = "#").place(x=50,y=60)

    global name
    name = tk.Entry(i)
    name.place(x=100, y=30)
    global num
    num = tk.Entry(i)
    num.place(x=100, y=60) 

    submitButton = tk.Button(i, text="Enter", command = submit).place(x=50,y=90)



playerImport = tk.Button(m, text='Import Players', width= 15, command = importPlayers)
playerImport.place(x=250,y=50)




m.mainloop()